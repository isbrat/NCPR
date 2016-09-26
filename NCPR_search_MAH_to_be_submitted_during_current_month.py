"""
Program to search in NCPR Pril 4 list for MAHs whose declarations should be submitted during the current month
Result is printed on the screen and exported to an Excel with 2 sheets: Products and Debug

NB: Downloaded NCPR file should be manipulated. Expected file:
- name: pril4towork.xlsx (or change on row 34)
- type: xlsx
- columns:
  1 - INN
  2 - Name
  3 - MAH
  4 - Note
  5 - Date of decision
  6 - Effective date

For now:
- MAH should be manually changed in 'searched MAH' (row 35)
- month should be manually changed in "current month' else it is the month from today (row 48)
"""

import openpyxl
from openpyxl.styles import Font
from datetime import datetime

# define columns
col_inn = 1
col_name = 2
col_MAH = 3
col_note = 4
col_decdate = 5
col_effdate = 6

# define files and MAH to search
file_NCPR = "pril4towork.xlsx"
searchedMAH = "Бакстер"
searched_months = []

# open file and load current sheet
file = openpyxl.load_workbook(file_NCPR)
sheet = file.active

#
today = datetime.now()
today_string = str(today)
month_today = today_string.split("-")
curr_month = month_today[1]
current_month = int(curr_month)
# current_month = 9

# get total number of rows in the sheet
row_count = int(sheet.max_row)

total_error_messages = 0

# make a variable for total number of products
total_products = 0

# create a new Excel file with name MAH+current month and make two sheets in it
name_for_export = searchedMAH + str(current_month) + ".xlsx"
new_file = openpyxl.Workbook()
first_sheet = new_file.get_active_sheet()
first_sheet.title = "Products"
second_sheet = new_file.create_sheet(title="Debug", index=1)

# main body

print("Starting to search for MAH " + searchedMAH +
      " to be submitted to NCPR until the end of month " + str(current_month) + " ... \n")

# find and exclude strikethrough rows
font = Font(strikethrough=None)
for i in range (3, row_count):
    MAH = sheet.cell(row=i, column=col_MAH).value
    current_font = sheet.cell(row=i, column=col_MAH).font.strike
    date_string = sheet.cell(row=i, column=col_effdate).value

# check if "Date" cell is empty and print a message + export to the Debug sheet in the output file
    if date_string == None:
        # print("     Debugging: No date on row " + str(i) + " - MAH is " + MAH)
        curr_cell = second_sheet["A" + str(total_error_messages + 1)]
        message = "No date on row " + str(i) + " - MAH is " + MAH
        curr_cell.value = message
        total_error_messages += 1
        pass
    else:
        # check if "date" is not a string but an integer - if integer skip and print a message
        # + export to the Debug sheet in the output file
        try:
            date_string == int(date_string)
            # print("     Debugging: Check date format on row " + str(i) + " - MAH is " + MAH)
            curr_cell = second_sheet["A" + str(total_error_messages + 1)]
            message = "Check date format on row " + str(i) + " - MAH is " + MAH
            curr_cell.value = message
            total_error_messages += 1
        except ValueError: #so date IS a string
            date_for_search = date_string.replace(",",".").rstrip(" г.").split(".")
            month = int(date_for_search[1])
            # prepare to check for current month and 6 months before
            if current_month < 7:
                searched_months = [current_month, current_month + 6]
            else:
                searched_months = [current_month, current_month - 6]
            if searchedMAH in MAH and current_font != True and month in searched_months: #print only non-strikethrough MAH to be submitted
                print("MAH " + searchedMAH + " exists on row " + str(i) +
                      " - and the product is: " + sheet.cell(row=i, column=col_name).value)
                curr_cell = first_sheet["A" + str(total_products + 1)]
                message = "MAH " + searchedMAH + " exists on row " + str(i) + \
                          " - and the product is: " + sheet.cell(row=i, column=col_name).value
                curr_cell.value = message
                total_products +=1
print("Total products to prepare: " + str(total_products))
new_file.save(name_for_export)
