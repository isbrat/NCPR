"""
Program to search in NCPR Pril 4 list for MAHs that are not striketrough

NB: Downloaded NCPR file should be manipulated. Expected file:
- name: pril4towork.xlsx (or change on row 32)
- type: xlsx
- columns:
  1 - INN
  2 - Name
  3 - MAH
  4 - Note
  5 - Date of decision
  6 - Effective date

For now:
- MAH should be manually changed in 'searched MAH' (row 33)
- month should be manually changed in "current month' else it is the month from today (row 46)
"""

import openpyxl
from openpyxl.styles import Font

# define columns
col_inn = 1
col_name = 2
col_MAH = 3
col_note = 4
col_decdate = 5
col_effdate = 6

# define files and MAH to search
file_NCPR = "pril4towork.xlsx"
searchedMAH = "Astellas"

# open file and load current sheet
file = openpyxl.load_workbook(file_NCPR)
sheet = file.active

# get total number of rows in the sheet
row_count = int(sheet.max_row)

# make a variable for total number of products
total_products = 0

# main body

print("Starting to search for MAH " + searchedMAH + " ... \n")

# find and exclude strikethrough rows
font = Font(strikethrough=None)
for i in range (3, row_count):
    MAH = sheet.cell(row=i, column=col_MAH).value
    current_font = sheet.cell(row=i, column=col_MAH).font.strike
    date_string = sheet.cell(row=i, column=col_effdate).value

# check for columns with MAH not strikethrough and print them
    if searchedMAH in MAH and current_font != True: # and month in searched_months: #print only non-strikethrough MAH
        print("MAH " + searchedMAH + " exists on row " + str(i) + " - "
              "and the product is: " + sheet.cell(row=i, column=col_name).value + ". "
              "\n     The effective date is " + sheet.cell(row=i, column=col_effdate).value)
        total_products +=1
print("Total products: " + str(total_products))
