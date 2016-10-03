"""
Program to search in NCPR Pril 4 list for MAHs that are not striketrough
NB: Downloaded NCPR file should be manipulated. Expected file:
- name: pril4towork.xlsx (or change on row 30)
- type: xlsx
- columns:
  1 - INN
  2 - Name
  3 - MAH
  4 - Note
  5 - Date of decision
  6 - Effective date
  7 - Last change date
- MAH should be manually changed in 'searched MAH' (row 31)
"""

import openpyxl

# define columns
col_inn = 1
col_name = 2
col_MAH = 3
col_note = 4
col_decdate = 5
col_effdate = 6
col_lastchangedate = 7

# define files and MAH to search
file_NCPR = "pril4towork.xlsx"
searchedMAH = "Бакстер"


def load_file():
    file = openpyxl.load_workbook(file_NCPR)
    return file


def get_active_sheet(file):
    active_sheet = file.active
    return active_sheet


def get_max_rows(active_sheet):
    max_rows = int(active_sheet.max_row)
    return max_rows


def exclude_strikethrough_rows_and_print():
    print("\n\nStart searching for MAH " + searchedMAH + " ... \n")
    total_products = 0
    for i in range(3, row_count):
        MAH = sheet.cell(row=i, column=col_MAH).value
        current_font = sheet.cell(row=i, column=col_MAH).font.strike

        # check for columns with MAH not strikethrough and print them
        if searchedMAH in MAH and current_font != True:  # print only non-strikethrough MAH
            print("MAH " + searchedMAH + " exists on row " + str(i) +
                  " - and the product is: \n " + sheet.cell(row=i, column=col_name).value + ". \n "
                  "The effective date is " + sheet.cell(row=i, column=col_effdate).value + ". \n "
                  "The date of last change is " + sheet.cell(row=i, column=col_lastchangedate).value)
            total_products += 1
    print("Total products: " + str(total_products))

work_file = load_file()
sheet = get_active_sheet(work_file)
row_count = get_max_rows(sheet)
exclude_strikethrough_rows_and_print()
