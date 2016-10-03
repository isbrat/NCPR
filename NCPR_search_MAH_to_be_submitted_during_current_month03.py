"""
Program to search in NCPR Pril 4 list for MAHs whose declarations should be submitted during the current month
Result is printed on the screen and exported to an Excel with 2 sheets: Products and Debug
NB: Downloaded NCPR file should be manipulated. Expected file:
- name: pril4towork.xlsx (or change on row 34)
- type: xlsx
- columns:
  1 - INN
  2 - Name
  3 - MAH
  4 - Note
  5 - Date of decision
  6 - Effective date
  7 - Last change date
For now:
- MAH should be manually changed in 'searched MAH' (row 35)
- month should be manually changed in "current month' else it is the month from today (row 48)
"""

import openpyxl
from openpyxl.styles import Font
from datetime import datetime

# define columns
col_inn = 1
col_name = 2
col_MAH = 3
col_note = 4
col_decdate = 5
col_effdate = 6
col_lastchangedate = 7

# define other variables
file_NCPR = "pril4towork.xlsx"
searchedMAH = "Астелас"
searched_months = []

total_error_messages = 0
total_products = 0


def get_current_month():
    today = datetime.now()
    today_string = str(today)
    month_today = today_string.split("-")
    curr_month = month_today[1]
    current_m = int(curr_month)
    # current_m = 11
    return current_m


def load_file():
    file = openpyxl.load_workbook(file_NCPR)
    return file


def get_active_sheet(file):
    active_sheet = file.active
    return active_sheet


def get_row_count(active_sheet):
    max_rows = int(sheet.max_row)
    return max_rows


def create_output_xlsx_file():
    # create a new Excel file with name MAH+current month and make two sheets in it
    output_file = openpyxl.Workbook()
    return output_file


def create_first_sheet(file):
    f_sheet = new_file.get_active_sheet()
    f_sheet.title = "Products"
    return f_sheet


def create_second_sheet(file):
    s_sheet = new_file.create_sheet(title="Debug", index=1)
    return s_sheet


def save_output_file(file, current_m):
    try:
        name_for_export = searchedMAH + str(current_month) + ".xlsx"
        new_file.save(name_for_export)
    except PermissionError as e:
        print("\nError saving the file - %s. Please close the file if it is open." % e)


def search_for_MAH(row_count, sheet, first_sheet, second_sheet, current_month, total_error_messages, total_products):
    print("\n\nStart searching for MAH " + searchedMAH +
          " to be submitted to NCPR until the end of month " + str(current_month) + " ... \n")
    font = Font(strikethrough=None)
    for i in range(3, row_count):
        MAH = sheet.cell(row=i, column=col_MAH).value
        current_font = sheet.cell(row=i, column=col_MAH).font.strike
        date_string = sheet.cell(row=i, column=col_effdate).value

        # check if "Date" cell is empty and print a message + export to the Debug sheet in the output file
        if date_string == None:
            # print("     Debugging: No date on row " + str(i) + " - MAH is " + MAH)
            curr_cell = second_sheet["A" + str(total_error_messages + 1)]
            message_MAH = "No date on row " + str(i) + " - MAH is " + MAH
            curr_cell.value = message_MAH
            total_error_messages += 1
            pass
        else:
            # check if "date" is not a string but an integer - if integer skip and print a message
            # + export to the Debug sheet in the output file
            try:
                date_string == int(date_string)
                curr_cell = second_sheet["A" + str(total_error_messages + 1)]
                message_MAH = "Check date format on row " + str(i) + " - MAH is " + MAH
                curr_cell.value = message_MAH
                total_error_messages += 1
            except ValueError:  # so date IS a string
                date_for_search = date_string.replace(",", ".").rstrip(" г.").split(".")
                month = int(date_for_search[1])
                # prepare to check for current month and 6 months before
                if current_month < 7:
                    searched_months = [current_month, current_month + 6]
                else:
                    searched_months = [current_month, current_month - 6]
                if searchedMAH in MAH and current_font != True and month in searched_months:  # print only non-strikethrough MAH to be submitted
                    print("MAH " + searchedMAH + " exists on row " + str(i) +
                          " - and the product is:\n " + sheet.cell(row=i, column=col_name).value +
                          "\nThe date of last change was " + sheet.cell(row=i, column=col_lastchangedate).value)
                    curr_cell = first_sheet["A" + str(total_products + 1)]
                    message_MAH = "MAH " + searchedMAH + " exists on row " + str(i) + \
                              " - and the product is: "
                    curr_cell.value = message_MAH
                    curr_cell = first_sheet["B" + str(total_products + 1)]
                    message_product = sheet.cell(row=i, column=col_name).value
                    curr_cell.value = message_product
                    curr_cell = first_sheet["C" + str(total_products + 1)]
                    message_lastchangedate = "Date of last change: " + sheet.cell(row=i, column=col_lastchangedate).value
                    curr_cell.value = message_lastchangedate
                    total_products += 1
    print("\nTotal products to prepare: " + str(total_products))

work_file = load_file()
sheet = get_active_sheet(work_file)
row_count = get_row_count(sheet)
current_month = get_current_month()
new_file = create_output_xlsx_file()
first_sheet = create_first_sheet(new_file)
second_sheet = create_second_sheet(new_file)
search_for_MAH(row_count, sheet, first_sheet, second_sheet, current_month, total_error_messages, total_products)
save_output_file(new_file,current_month)